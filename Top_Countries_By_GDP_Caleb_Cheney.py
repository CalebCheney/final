from itertools import count
from urllib.request import urlopen,Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

import re 
from openpyxl.descriptors import String, Sequence, Integer 
from openpyxl.descriptors.serialisable import Serialisable


BUILTIN_FORMATS = {
    0: 'General',
    1: '0',
    2: '0.00',
    3: '#,##0',
    4: '#,##0.00',
    5: '"$"#,##0_);("$"#,##0)',
    6: '"$"#,##0_);[Red]("$"#,##0)',
    7: '"$"#,##0.00_);("$"#,##0.00)',
    8: '"$"#,##0.00_);[Red]("$"#,##0.00)',
    9: '0%',
    10: '0.00%',
    11: '0.00E+00',
    12: '# ?/?',
    13: '# ??/??',
    14: 'mm-dd-yy',
    15: 'd-mmm-yy',
    16: 'd-mmm',
    17: 'mmm-yy',
    18: 'h:mm AM/PM',
    19: 'h:mm:ss AM/PM',
    20: 'h:mm',
    21: 'h:mm:ss',
    22: 'm/d/yy h:mm',

    37: '#,##0_);(#,##0)',
    38: '#,##0_);[Red](#,##0)',
    39: '#,##0.00_);(#,##0.00)',
    40: '#,##0.00_);[Red](#,##0.00)',

    41: r'_(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)',
    42: r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)',
    43: r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)',

    44: r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)',
    45: 'mm:ss',
    46: '[h]:mm:ss',
    47: 'mmss.0',
    48: '##0.0E+0',
    49: '@', }

BUILTIN_FORMATS_MAX_SIZE = 164
BUILTIN_FORMATS_REVERSE = dict(
        [(value, key) for key, value in BUILTIN_FORMATS.items()])

class NumberFormat(Serialisable):

    numFmtId = Integer()
    formatCode = String()

    def __init__(self,
                 numFmtId=None,
                 formatCode=None,
                ):
        self.numFmtId = numFmtId
        self.formatCode = formatCode

# scrape the website below to retrieve the top 5 countries with the highest GDPs. Calculate the GDP per capita
# by dividing the GDP by the population. You can perform the calculation in Python natively or insert the code
# in excel that will perform the calculation in Excel by each row. DO NOT scrape the GDP per capita from the
# webpage, make sure you use your own calculation.

# FOR YOUR REFERENCE - https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
# this link shows you the different number formats you can apply to a column using openpyxl


# FOR YOUR REFERENCE - https://www.geeksforgeeks.org/python-string-replace/
# this link shows you how to use the REPLACE function (you may need it if your code matches mine but not required)

### REMEMBER ##### - your output should match the excel file (GDP_Report.xlsx) including all formatting.

webpage = 'https://www.worldometers.info/gdp/gdp-by-country/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()			

soup = BeautifulSoup(webpage, 'html.parser')

title = soup.title
print(title.text)

stock_table = soup.findAll('tbody')

#gets first item in table
stock_table = stock_table[0]

#tr = table row
rows = stock_table.findAll('tr')

for row in rows[:5]:
    cols = row.findAll('td')
    num = cols[0]
    num = num.text
    country = cols[1]
    country = country.text
    gdp = cols[2]
    gdp = gdp.text.strip().strip('$').replace(',','')
    pop = cols[5]
    pop = pop.text.strip().replace(',','')
    gdp_cap = cols[6]
    gdp_cap = gdp_cap.text.strip('$').replace(',','')

'''
    print(num)
    print(country)
    print(gdp)
    print(pop)
    print(gdp_cap)
'''
wb = xl.Workbook()

MySheet = wb.active

MySheet.title = 'GDP By Country'

MySheet['A1'] = 'No.'
MySheet['B1'] = 'Country'
MySheet['C1'] = 'GDP'
MySheet['D1'] = 'Population'
MySheet['E1'] = 'GDP Per Capita'

MySheet['A1'].font = Font(size=16,bold=True)
MySheet['B1'].font = Font(size=16,bold=True)
MySheet['C1'].font = Font(size=16,bold=True)
MySheet['D1'].font = Font(size=16,bold=True)
MySheet['E1'].font = Font(size=16,bold=True)

for x in range(0, 5):
    cols = rows[x].findAll('td')
    num = cols[0]
    num = num.text
    country = cols[1]
    country = country.text
    gdp = cols[2]
    gdp = gdp.text.strip().strip('$').replace(',','')
    pop = cols[5]
    pop = pop.text.strip().replace(',','')
    gdp_cap = cols[6]
    gdp_cap = gdp_cap.text.strip().strip('$').replace(',','')

    print(num)
    print(country)
    print(gdp)
    print(pop)
    print(gdp_cap)

    

    MySheet['A' + str(x + 2)] = int(num)
    MySheet['B' + str(x + 2)] = country 
    MySheet['C' + str(x + 2)] = int(gdp) 
    MySheet['D' + str(x + 2)] = int(pop)
    MySheet['E' + str(x + 2)] = int(gdp_cap) 

MySheet.column_dimensions['A'].width = 5
MySheet.column_dimensions['B'].width = 15
MySheet.column_dimensions['C'].width = 25
MySheet.column_dimensions['D'].width = 18
MySheet.column_dimensions['E'].width = 20


#pop format
_cell = MySheet.cell(2,4)
_cell.number_format = '#,##0'

_cell = MySheet.cell(3,4)
_cell.number_format = '#,##0'

_cell = MySheet.cell(4,4)
_cell.number_format = '#,##0'

_cell = MySheet.cell(5,4)
_cell.number_format = '#,##0'

_cell = MySheet.cell(6,4)
_cell.number_format = '#,##0'

#gdp format
_cell = MySheet.cell(2,3)
_cell.number_format = '$#,##0_-'

_cell = MySheet.cell(3,3)
_cell.number_format = '$#,##0_-'

_cell = MySheet.cell(4,3)
_cell.number_format = '$#,##0_-'

_cell = MySheet.cell(5,3)
_cell.number_format = '$#,##0_-'

_cell = MySheet.cell(6,3)
_cell.number_format = '$#,##0_-'

#gdp_cap format
_cell = MySheet.cell(2,5)
_cell.number_format = '"$"#,##0.00_-'

_cell = MySheet.cell(3,5)
_cell.number_format = '"$"#,##0.00_-'

_cell = MySheet.cell(4,5)
_cell.number_format = '"$"#,##0.00_-'

_cell = MySheet.cell(5,5)
_cell.number_format = '"$"#,##0.00_-'

_cell = MySheet.cell(6,5)
_cell.number_format = '"$"#,##0.00_-'

wb.save('GDP_Report_Caleb_Cheney.xlsx')
    





